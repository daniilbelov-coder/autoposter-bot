"""
Модуль для генерации еженедельного календаря коммуникаций в Excel.
Использует умную логику распределения постов с учетом частоты и конфликтов.
"""

import os
import random
from datetime import datetime, timedelta
from typing import List, Dict, Set
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import POSTING_TIMES, TIMEZONE, logger
from message_selector import selector
from database import db


def get_week_dates(start_date: datetime = None) -> List[datetime]:
    """
    Получить даты недели (понедельник-воскресенье).
    
    Args:
        start_date: Начальная дата (если None - текущий понедельник)
        
    Returns:
        Список дат недели
    """
    if start_date is None:
        today = datetime.now(TIMEZONE)
        # Получить понедельник текущей недели
        start_date = today - timedelta(days=today.weekday())
    
    week_dates = []
    for i in range(7):
        week_dates.append(start_date + timedelta(days=i))
    
    return week_dates


def can_add_post_to_day(
    schedule: Dict,
    date: datetime,
    message_id: int,
    message_title: str,
    conflict_ids: List[int]
) -> bool:
    """
    Проверить, можно ли добавить пост в указанный день.
    
    Args:
        schedule: Расписание (ключ - дата, значение - список постов)
        date: Дата для проверки
        message_id: ID сообщения
        message_title: Название сообщения
        conflict_ids: Список ID сообщений, с которыми нельзя публиковаться в один день
        
    Returns:
        True, если пост можно добавить
    """
    date_key = date.date()
    
    # Проверяем посты в этот день
    if date_key in schedule:
        existing_posts = schedule[date_key]
        
        # Проверка 1: Этот пост уже запланирован на этот день
        if any(p['message_id'] == message_id for p in existing_posts):
            return False
        
        # Проверка 2: Конфликтующие посты в этот день
        existing_ids = [p['message_id'] for p in existing_posts]
        if any(conflict_id in existing_ids for conflict_id in conflict_ids):
            return False
    
    # Проверяем соседние дни (2 дня до и 2 после)
    for days_offset in range(-2, 3):
        if days_offset == 0:
            continue  # Текущий день уже проверили
        
        check_date = (date + timedelta(days=days_offset)).date()
        
        if check_date in schedule:
            # Проверяем, есть ли этот же пост в соседних днях
            if any(p['message_id'] == message_id for p in schedule[check_date]):
                return False
    
    return True


def get_day_post_count(schedule: Dict, date: datetime) -> int:
    """Получить количество постов в указанный день."""
    date_key = date.date()
    return len(schedule.get(date_key, []))


def calculate_message_priority(message: Dict, current_date: datetime) -> float:
    """
    Рассчитать приоритет сообщения для публикации.
    Чем выше приоритет, тем важнее опубликовать сообщение.
    
    Args:
        message: Словарь с данными сообщения
        current_date: Текущая дата планирования
        
    Returns:
        Приоритет (число от 0 до 100)
    """
    message_id = message['id']
    frequency_days = message['frequency_days']
    
    # Получить дату последней отправки
    last_sent_date = db.get_last_sent_date(message_id)
    
    if last_sent_date is None:
        # Сообщение никогда не отправлялось - высокий приоритет
        return 100.0
    
    # Рассчитать, сколько дней прошло
    days_passed = (current_date.date() - last_sent_date).days
    
    # Рассчитать приоритет на основе прошедшего времени
    if days_passed < frequency_days:
        # Еще рано публиковать
        return 0.0
    elif days_passed == frequency_days:
        # Идеальное время
        return 50.0
    else:
        # Чем больше прошло времени, тем выше приоритет
        overdue_days = days_passed - frequency_days
        priority = 50.0 + min(overdue_days * 5, 50.0)  # До 100
        return priority


def get_planned_posts_for_week_smart() -> List[Dict]:
    """
    Получить список запланированных постов на неделю.
    Использует умную логику распределения с учетом частоты и конфликтов.
    
    Returns:
        Список словарей с информацией о постах
    """
    week_dates = get_week_dates()
    posts_per_day = len(POSTING_TIMES)
    
    # Создаем расписание: ключ - дата, значение - список постов
    schedule: Dict[datetime.date, List[Dict]] = {}
    
    # Получаем все сообщения
    all_messages = selector.messages
    
    if not all_messages:
        logger.warning("Нет сообщений для планирования")
        return []
    
    logger.info(f"Начало планирования на неделю: {len(all_messages)} сообщений доступно")
    
    # Инициализируем расписание пустыми списками
    for date in week_dates:
        schedule[date.date()] = []
    
    # Рассчитываем приоритеты сообщений
    message_priorities = []
    for message in all_messages:
        priority = calculate_message_priority(message, week_dates[0])
        if priority > 0:  # Только сообщения, которые уже можно публиковать
            message_priorities.append({
                'message': message,
                'priority': priority
            })
    
    # Сортируем по приоритету (от высокого к низкому)
    message_priorities.sort(key=lambda x: x['priority'], reverse=True)
    
    logger.info(f"Сообщений с приоритетом > 0: {len(message_priorities)}")
    
    # Подсчитываем, сколько раз каждое сообщение должно быть опубликовано
    # На основе frequency_days определяем частоту публикации в неделю
    message_weekly_frequency = {}
    for msg_data in message_priorities:
        message = msg_data['message']
        freq_days = message['frequency_days']
        
        # Рассчитываем частоту публикации в неделю
        if freq_days <= 3:  # Ежедневно или несколько раз в неделю
            weekly_freq = 7 // freq_days
        elif freq_days == 7:  # Раз в неделю
            weekly_freq = 1
        elif freq_days == 14:  # Раз в 2 недели
            weekly_freq = 1  # В одну из двух недель
        elif freq_days == 21:  # Раз в 3 недели
            weekly_freq = 1  # В одну из трех недель
        else:  # Месяц и больше
            weekly_freq = 0  # Не в каждую неделю
        
        message_weekly_frequency[message['id']] = {
            'target': weekly_freq,
            'placed': 0
        }
    
    # Распределяем посты по дням
    # Проход 1: Размещаем высокоприоритетные посты
    for msg_data in message_priorities:
        message = msg_data['message']
        message_id = message['id']
        target_freq = message_weekly_frequency[message_id]['target']
        
        if target_freq == 0:
            continue  # Это сообщение не нужно публиковать на этой неделе
        
        # Пытаемся разместить сообщение target_freq раз
        placed_count = 0
        attempts = 0
        max_attempts = 50
        
        while placed_count < target_freq and attempts < max_attempts:
            attempts += 1
            
            # Ищем день с минимальной загрузкой, куда можно добавить пост
            best_date = None
            min_posts = float('inf')
            
            # Создаем список дней с их текущей загрузкой
            day_loads = []
            for date in week_dates:
                post_count = get_day_post_count(schedule, date)
                
                # Проверяем, можно ли добавить пост в этот день
                if post_count < posts_per_day:
                    if can_add_post_to_day(
                        schedule,
                        date,
                        message_id,
                        message['title'],
                        message.get('do_not_schedule_same_day_with', [])
                    ):
                        day_loads.append((date, post_count))
            
            if not day_loads:
                # Не удалось найти подходящий день
                logger.warning(f"Не удалось разместить сообщение {message_id} ({message['title']})")
                break
            
            # Выбираем день с минимальной загрузкой
            day_loads.sort(key=lambda x: x[1])
            best_date = day_loads[0][0]
            
            # Если несколько дней с одинаковой загрузкой, выбираем случайно
            min_load = day_loads[0][1]
            days_with_min_load = [d for d, l in day_loads if l == min_load]
            best_date = random.choice(days_with_min_load)
            
            # Добавляем пост в этот день
            schedule[best_date.date()].append({
                'message_id': message_id,
                'message': message
            })
            
            placed_count += 1
            message_weekly_frequency[message_id]['placed'] += 1
            
            logger.debug(f"Размещено: {message['title']} на {best_date.date()}, "
                        f"попытка {attempts}, размещено {placed_count}/{target_freq}")
        
        if placed_count < target_freq:
            logger.warning(f"Сообщение {message_id} размещено {placed_count} раз вместо {target_freq}")
    
    # Проход 2: Заполняем пустые слоты случайными сообщениями из доступных
    for date in week_dates:
        current_posts = get_day_post_count(schedule, date)
        
        if current_posts < posts_per_day:
            # Нужно добавить еще постов
            slots_to_fill = posts_per_day - current_posts
            
            # Ищем сообщения, которые можно добавить
            for _ in range(slots_to_fill):
                # Перемешиваем сообщения для разнообразия
                random_messages = list(all_messages)
                random.shuffle(random_messages)
                
                added = False
                for message in random_messages:
                    if can_add_post_to_day(
                        schedule,
                        date,
                        message['id'],
                        message['title'],
                        message.get('do_not_schedule_same_day_with', [])
                    ):
                        schedule[date.date()].append({
                            'message_id': message['id'],
                            'message': message
                        })
                        added = True
                        logger.debug(f"Дополнительно размещено: {message['title']} на {date.date()}")
                        break
                
                if not added:
                    logger.warning(f"Не удалось заполнить слот в день {date.date()}")
                    # Добавляем пустой слот
                    schedule[date.date()].append({
                        'message_id': None,
                        'message': None
                    })
    
    # Преобразуем расписание в список постов с временными слотами
    planned_posts = []
    
    for date in week_dates:
        date_posts = schedule[date.date()]
        
        # Перемешиваем посты в дне для разнообразия времени
        random.shuffle(date_posts)
        
        for idx, post_data in enumerate(date_posts):
            if idx >= len(POSTING_TIMES):
                break
            
            time_config = POSTING_TIMES[idx]
            post_datetime = date.replace(
                hour=time_config['hour'],
                minute=time_config['minute'],
                second=0,
                microsecond=0
            )
            
            message = post_data.get('message')
            
            if message:
                planned_posts.append({
                    'date': post_datetime.date(),
                    'time': post_datetime.time(),
                    'datetime': post_datetime,
                    'day_name': post_datetime.strftime('%A'),
                    'message_id': message['id'],
                    'message_title': message['title'],
                    'frequency': message.get('frequency', 'unknown'),
                    'has_photos': len(message.get('photos', [])) > 0,
                    'photos_count': len(message.get('photos', [])),
                    'has_videos': len(message.get('videos', [])) > 0,
                    'videos_count': len(message.get('videos', []))
                })
            else:
                # Пустой слот
                planned_posts.append({
                    'date': post_datetime.date(),
                    'time': post_datetime.time(),
                    'datetime': post_datetime,
                    'day_name': post_datetime.strftime('%A'),
                    'message_id': None,
                    'message_title': '—',
                    'frequency': '',
                    'has_photos': False,
                    'photos_count': 0,
                    'has_videos': False,
                    'videos_count': 0
                })
    
    # Сортируем по дате и времени
    planned_posts.sort(key=lambda x: x['datetime'])
    
    logger.info(f"Запланировано постов: {len([p for p in planned_posts if p['message_id']])}")
    
    return planned_posts


def create_schedule_excel(output_path: str = "schedule.xlsx") -> str:
    """
    Создать Excel файл с календарем коммуникаций.
    
    Args:
        output_path: Путь для сохранения файла
        
    Returns:
        Путь к созданному файлу
    """
    logger.info("Начало генерации календаря коммуникаций")
    
    # Создать книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Календарь коммуникаций"
    
    # Получить даты недели
    week_dates = get_week_dates()
    start_date = week_dates[0]
    end_date = week_dates[-1]
    
    # Получить запланированные посты (умная логика)
    planned_posts = get_planned_posts_for_week_smart()
    
    # Стили
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='Arial', size=10)
    
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    alternate_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Заголовок документа
    ws.merge_cells('A1:G1')
    ws['A1'] = f'КАЛЕНДАРЬ КОММУНИКАЦИЙ'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 25
    
    # Подзаголовок с датами
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Период: {start_date.strftime("%d.%m.%Y")} - {end_date.strftime("%d.%m.%Y")}'
    ws['A2'].font = Font(name='Arial', size=11, italic=True)
    ws['A2'].alignment = center_alignment
    ws.row_dimensions[2].height = 20
    
    # Пустая строка
    ws.row_dimensions[3].height = 5
    
    # Заголовки столбцов
    headers = ['№', 'Дата', 'День недели', 'Время', 'Название сообщения', 'Частота', 'Фото']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    ws.row_dimensions[4].height = 20
    
    # Данные
    day_names_ru = {
        'Monday': 'Понедельник',
        'Tuesday': 'Вторник',
        'Wednesday': 'Среда',
        'Thursday': 'Четверг',
        'Friday': 'Пятница',
        'Saturday': 'Суббота',
        'Sunday': 'Воскресенье'
    }
    
    # Фильтруем только заполненные слоты
    filled_posts = [p for p in planned_posts if p['message_id'] is not None]
    
    for idx, post in enumerate(filled_posts, 1):
        row = idx + 4
        
        # Чередующийся цвет строк
        fill = alternate_fill if idx % 2 == 0 else PatternFill()
        
        # Номер
        cell = ws.cell(row=row, column=1, value=idx)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Дата
        cell = ws.cell(row=row, column=2, value=post['date'].strftime('%d.%m.%Y'))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # День недели
        day_name = day_names_ru.get(post['day_name'], post['day_name'])
        cell = ws.cell(row=row, column=3, value=day_name)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Время
        cell = ws.cell(row=row, column=4, value=post['time'].strftime('%H:%M'))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Название сообщения
        cell = ws.cell(row=row, column=5, value=post['message_title'])
        cell.font = cell_font
        cell.alignment = left_alignment
        cell.border = border
        cell.fill = fill
        
        # Частота
        frequency_ru = {
            'каждые 2 недели': 'каждые 2 недели',
            'раз в месяц': 'раз в месяц',
            'раз в неделю': 'раз в неделю',
            'раз в две недели': 'раз в две недели',
            'раз в 3 недели': 'раз в 3 недели',
        }
        freq_text = frequency_ru.get(post['frequency'], post['frequency'])
        cell = ws.cell(row=row, column=6, value=freq_text)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Фото/Видео
        media_parts = []
        if post['has_photos']:
            media_parts.append(f"📷 ({post['photos_count']})")
        if post['has_videos']:
            media_parts.append(f"📹 ({post['videos_count']})")
        
        photos_text = " ".join(media_parts) if media_parts else "—"
        cell = ws.cell(row=row, column=7, value=photos_text)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        ws.row_dimensions[row].height = 18
    
    # Футер
    footer_row = len(filled_posts) + 6
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    ws[f'A{footer_row}'] = f'Всего запланировано публикаций: {len(filled_posts)}'
    ws[f'A{footer_row}'].font = Font(name='Arial', size=10, italic=True)
    ws[f'A{footer_row}'].alignment = center_alignment
    
    # Подпись
    footer_row += 1
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    ws[f'A{footer_row}'] = f'Сгенерировано: {datetime.now(TIMEZONE).strftime("%d.%m.%Y %H:%M")}'
    ws[f'A{footer_row}'].font = Font(name='Arial', size=9, italic=True, color='808080')
    ws[f'A{footer_row}'].alignment = center_alignment
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 12
    
    # Сохранить файл
    wb.save(output_path)
    logger.info(f"Календарь сохранён: {output_path}")
    
    return output_path


def generate_schedule_for_subscribers() -> str:
    """
    Сгенерировать календарь для рассылки подписчикам.
    
    Returns:
        Путь к файлу
    """
    # Создать папку для расписаний, если нет
    schedules_dir = "schedules"
    if not os.path.exists(schedules_dir):
        os.makedirs(schedules_dir)
    
    # Имя файла с датой
    filename = f"calendar_{datetime.now(TIMEZONE).strftime('%Y_%m_%d')}.xlsx"
    filepath = os.path.join(schedules_dir, filename)
    
    return create_schedule_excel(filepath)


if __name__ == '__main__':
    # Тестовая генерация
    print("Генерация тестового календаря...")
    filepath = generate_schedule_for_subscribers()
    print(f"✅ Календарь создан: {filepath}")
