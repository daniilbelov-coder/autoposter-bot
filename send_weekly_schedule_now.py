"""
Скрипт для немедленной отправки еженедельного расписания.
Генерирует календарь с текущего момента до конца недели.
"""

import asyncio
import os
from datetime import datetime, timedelta
from telegram import Bot
from config import BOT_TOKEN, TIMEZONE, logger
from database import db
from schedule_generator import create_schedule_excel, get_week_dates
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from message_selector import selector
from config import POSTING_TIMES


def get_remaining_week_dates() -> list:
    """
    Получить даты с текущего дня до конца недели (воскресенье).
    
    Returns:
        Список дат
    """
    today = datetime.now(TIMEZONE)
    current_weekday = today.weekday()  # 0 = понедельник, 6 = воскресенье
    
    remaining_dates = []
    
    # Дни до воскресенья включительно
    days_until_sunday = 6 - current_weekday
    
    for i in range(days_until_sunday + 1):
        date = today + timedelta(days=i)
        remaining_dates.append(date)
    
    return remaining_dates


def get_planned_posts_for_remaining_week() -> list:
    """
    Получить список постов с текущего момента до конца недели.
    
    Returns:
        Список словарей с информацией о постах
    """
    remaining_dates = get_remaining_week_dates()
    current_time = datetime.now(TIMEZONE)
    planned_posts = []
    
    # Для каждого дня
    for date in remaining_dates:
        # Для каждого времени отправки
        for time_config in POSTING_TIMES:
            hour = time_config['hour']
            minute = time_config['minute']
            
            post_datetime = date.replace(hour=hour, minute=minute, second=0, microsecond=0)
            
            # Пропустить прошедшее время сегодня
            if post_datetime < current_time:
                continue
            
            # Получить случайное сообщение для примера
            available_messages = selector.messages
            
            if available_messages:
                import random
                message = random.choice(available_messages)
                
                planned_posts.append({
                    'date': post_datetime.date(),
                    'time': post_datetime.time(),
                    'datetime': post_datetime,
                    'day_name': post_datetime.strftime('%A'),
                    'message_id': message['id'],
                    'message_title': message['title'],
                    'frequency': message['frequency'],
                    'has_photos': len(message.get('photos', [])) > 0,
                    'photos_count': len(message.get('photos', []))
                })
    
    # Сортировать по дате и времени
    planned_posts.sort(key=lambda x: x['datetime'])
    
    return planned_posts


def create_remaining_week_schedule(output_path: str = "schedule_current_week.xlsx") -> str:
    """
    Создать календарь для оставшейся части недели.
    
    Args:
        output_path: Путь для сохранения
        
    Returns:
        Путь к файлу
    """
    logger.info("Генерация календаря для оставшейся части недели")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание до конца недели"
    
    # Получить даты
    remaining_dates = get_remaining_week_dates()
    start_date = remaining_dates[0]
    end_date = remaining_dates[-1]
    
    # Получить запланированные посты
    planned_posts = get_planned_posts_for_remaining_week()
    
    if not planned_posts:
        logger.warning("Нет запланированных постов до конца недели!")
        # Все равно создадим файл с пустой таблицей
    
    # Стили
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='Arial', size=10)
    
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    alternate_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Заголовок
    ws.merge_cells('A1:G1')
    current_time = datetime.now(TIMEZONE)
    ws['A1'] = f'РАСПИСАНИЕ ПУБЛИКАЦИЙ ДО КОНЦА НЕДЕЛИ'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 25
    
    # Подзаголовок
    ws.merge_cells('A2:G2')
    ws['A2'] = f'С {start_date.strftime("%d.%m.%Y %H:%M")} по {end_date.strftime("%d.%m.%Y")} 23:59'
    ws['A2'].font = Font(name='Arial', size=11, italic=True)
    ws['A2'].alignment = center_alignment
    ws.row_dimensions[2].height = 20
    
    # Пустая строка
    ws.row_dimensions[3].height = 5
    
    # Заголовки столбцов
    headers = ['№', 'Дата', 'День недели', 'Время', 'Название сообщения', 'Частота', 'Фото']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    ws.row_dimensions[4].height = 20
    
    # Данные
    day_names_ru = {
        'Monday': 'Понедельник',
        'Tuesday': 'Вторник',
        'Wednesday': 'Среда',
        'Thursday': 'Четверг',
        'Friday': 'Пятница',
        'Saturday': 'Суббота',
        'Sunday': 'Воскресенье'
    }
    
    for idx, post in enumerate(planned_posts, 1):
        row = idx + 4
        
        fill = alternate_fill if idx % 2 == 0 else PatternFill()
        
        # Номер
        cell = ws.cell(row=row, column=1, value=idx)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Дата
        cell = ws.cell(row=row, column=2, value=post['date'].strftime('%d.%m.%Y'))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # День недели
        day_name = day_names_ru.get(post['day_name'], post['day_name'])
        cell = ws.cell(row=row, column=3, value=day_name)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Время
        cell = ws.cell(row=row, column=4, value=post['time'].strftime('%H:%M'))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Название
        cell = ws.cell(row=row, column=5, value=post['message_title'])
        cell.font = cell_font
        cell.alignment = left_alignment
        cell.border = border
        cell.fill = fill
        
        # Частота
        frequency_ru = {
            'daily': 'Ежедневно',
            'weekly': 'Еженедельно',
            'biweekly': 'Раз в 2 недели',
            'monthly': 'Ежемесячно'
        }
        freq_text = frequency_ru.get(post['frequency'], post['frequency'])
        cell = ws.cell(row=row, column=6, value=freq_text)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Фото
        photos_text = f"✓ ({post['photos_count']})" if post['has_photos'] else "—"
        cell = ws.cell(row=row, column=7, value=photos_text)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        ws.row_dimensions[row].height = 18
    
    # Футер
    footer_row = len(planned_posts) + 6
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    ws[f'A{footer_row}'] = f'Запланировано публикаций до конца недели: {len(planned_posts)}'
    ws[f'A{footer_row}'].font = Font(name='Arial', size=10, bold=True)
    ws[f'A{footer_row}'].alignment = center_alignment
    
    # Подпись
    footer_row += 1
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    ws[f'A{footer_row}'] = f'Сгенерировано: {current_time.strftime("%d.%m.%Y %H:%M")}'
    ws[f'A{footer_row}'].font = Font(name='Arial', size=9, italic=True, color='808080')
    ws[f'A{footer_row}'].alignment = center_alignment
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10
    
    # Сохранить
    wb.save(output_path)
    logger.info(f"Календарь сохранён: {output_path}")
    
    return output_path


async def send_schedule_now(test_user_id: int = None):
    """
    Отправить расписание всем подписчикам прямо сейчас.
    
    Args:
        test_user_id: ID пользователя для тестовой отправки (опционально)
    """
    print("=" * 70)
    print("📅 ОТПРАВКА ЕЖЕНЕДЕЛЬНОГО РАСПИСАНИЯ")
    print("=" * 70)
    
    # Создать папку schedules если нет
    if not os.path.exists('schedules'):
        os.makedirs('schedules')
    
    # Сгенерировать календарь
    print("\n⏳ Генерация календаря...")
    schedule_file = create_remaining_week_schedule('schedules/schedule_current_week.xlsx')
    print(f"✅ Календарь создан: {schedule_file}")
    
    # Получить подписчиков
    print("\n⏳ Получение списка подписчиков...")
    subscribers = db.get_active_subscribers()
    
    # Если указан test_user_id, отправить только ему
    if test_user_id:
        print(f"🧪 Тестовый режим: отправка только пользователю {test_user_id}")
        subscribers = [{'user_id': test_user_id, 'username': 'test', 'first_name': 'Test'}]
    elif not subscribers:
        print("⚠️  Нет активных подписчиков!")
        print("\n💡 Запустите скрипт с вашим User ID:")
        print("   python send_weekly_schedule_now.py YOUR_USER_ID")
        print("\n   Узнать User ID: отправьте любое сообщение боту @userinfobot")
        return
    
    print(f"✅ Получателей: {len(subscribers)}")
    
    # Создать бота
    bot = Bot(token=BOT_TOKEN)
    
    # Текст сообщения
    message_text = (
        "📅 <b>Расписание публикаций до конца недели</b>\n\n"
        "Здравствуйте! Высылаем вам план публикаций на оставшуюся часть недели.\n\n"
        "В файле вы найдете:\n"
        "• Даты и время всех запланированных публикаций\n"
        "• Названия сообщений\n"
        "• Информацию о фотографиях\n\n"
        "📊 Это тестовая отправка расписания!"
    )
    
    # Отправить каждому
    print("\n📤 Начало рассылки...\n")
    success_count = 0
    error_count = 0
    
    for subscriber in subscribers:
        user_id = subscriber['user_id']
        username = subscriber.get('username', 'unknown')
        
        try:
            with open(schedule_file, 'rb') as file:
                await bot.send_document(
                    chat_id=user_id,
                    document=file,
                    caption=message_text,
                    parse_mode='HTML',
                    filename='schedule_remaining_week.xlsx'
                )
            
            print(f"✅ Отправлено пользователю {user_id} (@{username})")
            success_count += 1
            
            # Обновить время последней отправки
            db.update_last_sent_schedule(user_id)
            
        except Exception as e:
            print(f"❌ Ошибка для {user_id} (@{username}): {e}")
            error_count += 1
        
        # Задержка между отправками
        await asyncio.sleep(0.5)
    
    print("\n" + "=" * 70)
    print("📊 РЕЗУЛЬТАТЫ РАССЫЛКИ")
    print("=" * 70)
    print(f"✅ Успешно: {success_count}")
    print(f"❌ Ошибок: {error_count}")
    print(f"📊 Всего: {len(subscribers)}")
    print("=" * 70)
    
    if success_count > 0:
        print("\n🎉 Рассылка завершена! Проверьте Telegram.")
    else:
        print("\n⚠️  Ни одно сообщение не отправлено.")


if __name__ == '__main__':
    import sys
    
    print("\n🤖 Скрипт отправки еженедельного расписания\n")
    print("Сегодня:", datetime.now(TIMEZONE).strftime("%A, %d.%m.%Y %H:%M"))
    print("\n")
    
    # Проверить аргументы командной строки
    test_user_id = None
    if len(sys.argv) > 1:
        try:
            test_user_id = int(sys.argv[1])
            print(f"📝 Тестовый режим для User ID: {test_user_id}\n")
        except ValueError:
            print("❌ Ошибка: User ID должен быть числом")
            print("Использование: python send_weekly_schedule_now.py YOUR_USER_ID")
            sys.exit(1)
    
    try:
        asyncio.run(send_schedule_now(test_user_id))
    except KeyboardInterrupt:
        print("\n\n❌ Отменено пользователем")
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        logger.error(f"Ошибка в скрипте рассылки: {e}", exc_info=True)
