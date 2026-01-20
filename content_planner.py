import pandas as pd
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os

def read_input_file(file_path):
    """Чтение входного Excel файла"""
    try:
        df = pd.read_excel(file_path)
        if len(df.columns) < 4:
            raise ValueError("Файл должен содержать как минимум 4 столбца")
        
        # Создаем новый DataFrame с нужными названиями столбцов
        new_df = pd.DataFrame()
        
        # Копируем данные из первых 4 столбцов
        new_df['A'] = df.iloc[:, 0]  # Название поста
        new_df['B'] = df.iloc[:, 1]  # Количество в неделю
        new_df['C'] = df.iloc[:, 2]  # Количество в месяц
        new_df['D'] = df.iloc[:, 3]  # Ссылка
        
        # Добавляем столбец E (чередование), если есть 5-й столбец
        if len(df.columns) >= 5:
            new_df['E'] = df.iloc[:, 4]  # Чередование
        else:
            new_df['E'] = None
        
        return new_df
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        return None

def validate_input(df):
    """Проверка корректности входных данных"""
    for _, row in df.iterrows():
        if pd.notna(row['B']) and pd.notna(row['C']):
            return False, "В одной строке не могут быть заполнены оба столбца B и C"
        
        # Проверяем корректность чередования
        if pd.notna(row['E']):
            try:
                # Проверяем, что все указанные номера строк существуют
                alternate_rows = [int(x.strip()) for x in str(row['E']).split(',')]
                if any(x < 2 or x > len(df) + 1 for x in alternate_rows):
                    return False, f"Некорректный номер строки в столбце E для поста '{row['A']}'"
            except ValueError:
                return False, f"Некорректный формат в столбце E для поста '{row['A']}'"
    
    return True, ""

def can_add_post(all_days, week_idx, day_idx, post_title, is_monthly=False):
    """Проверяет, можно ли добавить пост в указанный день"""
    current_day = all_days[week_idx][day_idx]['date']
    
    # Проверяем этот день
    if any(post['title'] == post_title for post in all_days[week_idx][day_idx]['posts']):
        return False
    
    # Проверяем соседние дни (2 дня до и 2 дня после)
    for w in range(max(0, week_idx-1), min(len(all_days), week_idx+2)):
        for d in range(7):
            # Вычисляем разницу в днях
            check_day = all_days[w][d]['date']
            days_diff = abs((check_day - current_day).days)
            
            # Проверяем интервал в 2 дня
            if days_diff <= 2:
                if any(post['title'] == post_title for post in all_days[w][d]['posts']):
                    return False
    
    if is_monthly:
        # Для ежемесячных постов проверяем соседние недели
        for w in range(max(0, week_idx-1), min(len(all_days), week_idx+2)):
            for d in range(7):
                if any(post['title'] == post_title for post in all_days[w][d]['posts']):
                    return False
    
    return True

def count_posts_in_range(all_days, week_idx, day_idx, range_days=2):
    """Подсчитывает количество постов в указанном диапазоне дней"""
    count = 0
    current_day = all_days[week_idx][day_idx]['date']
    
    for w in range(max(0, week_idx-1), min(len(all_days), week_idx+2)):
        for d in range(7):
            check_day = all_days[w][d]['date']
            days_diff = abs((check_day - current_day).days)
            if days_diff <= range_days:
                count += len([p for p in all_days[w][d]['posts'] if p['title']])
    
    return count

def find_best_day(all_days, week_idx, possible_days, post_title, is_monthly=False):
    """Находит лучший день для размещения поста"""
    min_posts = float('inf')
    best_days = []
    
    for day_idx in possible_days:
        if can_add_post(all_days, week_idx, day_idx, post_title, is_monthly):
            posts_count = count_posts_in_range(all_days, week_idx, day_idx)
            if posts_count < min_posts:
                min_posts = posts_count
                best_days = [day_idx]
            elif posts_count == min_posts:
                best_days.append(day_idx)
    
    return random.choice(best_days) if best_days else None

def get_post_schedule(df, weeks, posts_per_day, start_date=None):
    """Создание расписания постов"""
    post_slots = ['08:00', '10:00', '12:00', '14:00', '16:00']
    
    # Создаем список всех постов с их частотой
    weekly_posts = []  # Посты, которые нужно публиковать раз в неделю
    monthly_posts = []  # Посты, которые нужно публиковать раз в месяц
    
    # Создаем группы чередования
    alternation_groups = {}
    post_to_group = {}  # Соответствие между постом и его группой
    
    # Первый проход: создаем посты и определяем группы чередования
    for idx, row in df.iterrows():
        # Номер строки в Excel (учитывая, что первая строка - заголовки)
        row_number = idx + 2
        
        post_info = {
            'title': row['A'],
            'link': row['D'],
            'row_number': row_number,
            'in_alternation_group': False
        }
        
        if pd.notna(row['B']):  # Посты в неделю
            frequency = float(row['B'])
            if frequency == 0.5:
                # Для постов раз в 2 недели
                post_info.update({
                    'frequency': 1,
                    'interval_weeks': 2,
                    'placed_in_week': [0] * (weeks + 1)
                })
            else:
                post_info.update({
                    'frequency': int(frequency),
                    'interval_weeks': 1,
                    'placed_in_week': [0] * weeks
                })
            weekly_posts.append(post_info)
            
        elif pd.notna(row['C']):  # Посты в месяц
            frequency = float(row['C'])
            if frequency == 0.5:
                # Для постов раз в 2 месяца
                post_info.update({
                    'frequency': 1,
                    'interval_months': 2,
                    'placed_in_month': [0] * (weeks // 8 + 1)
                })
            else:
                post_info.update({
                    'frequency': int(frequency),
                    'interval_months': 1,
                    'placed_in_month': [0] * (weeks // 4 + 1)
                })
            monthly_posts.append(post_info)
        
        # Если у поста есть указание на чередование
        if pd.notna(row['E']):
            alternate_rows = [int(x.strip()) for x in str(row['E']).split(',')]
            group = [row_number] + alternate_rows
            group_key = tuple(sorted(group))
            
            if group_key not in alternation_groups:
                is_weekly = pd.notna(row['B'])
                frequency_value = float(row['B']) if is_weekly else float(row['C'])
                interval = 1
                
                # Определяем интервал для постов с частотой 0.5
                if frequency_value == 0.5:
                    interval = 2
                    frequency_value = 1
                
                alternation_groups[group_key] = {
                    'posts': [],
                    'current_index': 0,
                    'weekly': is_weekly,
                    'frequency': int(frequency_value),
                    'interval_weeks': interval if is_weekly else 1,
                    'interval_months': interval if not is_weekly else 1
                }
    
    # Второй проход: заполняем группы чередования
    for post in weekly_posts + monthly_posts:
        for group_key, group in alternation_groups.items():
            if post['row_number'] in group_key:
                group['posts'].append(post)
                post['in_alternation_group'] = True
                post_to_group[post['row_number']] = group_key
    
    # Создаем расписание
    current_date = start_date if start_date else datetime.now()
    current_date = current_date.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Создаем все дни периода
    all_days = []
    for week in range(weeks):
        week_days = []
        for day in range(7):
            current_day = current_date + timedelta(days=day + week*7)
            week_days.append({
                'date': current_day,
                'posts': []
            })
        all_days.append(week_days)
    
    def get_day_load(week_idx, day_idx):
        """Вычисляет загруженность дня"""
        return len([p for p in all_days[week_idx][day_idx]['posts'] if p['title']])
    
    def get_week_load(week_idx):
        """Вычисляет загруженность недели"""
        return sum(get_day_load(week_idx, d) for d in range(7))
    
    # Распределяем посты по неделям с учетом чередования
    
    # Обрабатываем еженедельные посты
    for week_idx in range(weeks):
        # Обрабатываем группы чередования для еженедельных постов
        for group_key, group in alternation_groups.items():
            if group['weekly']:  # Только для еженедельных групп
                # Пропускаем, если интервал 2 недели и это нечетная неделя
                if 'interval_weeks' in group and group['interval_weeks'] == 2 and week_idx % 2 != 0:
                    continue
                    
                frequency = group['frequency']
                
                # Определяем, какой пост из группы выходит в эту неделю
                post_index = (week_idx // group.get('interval_weeks', 1) % len(group['posts']))
                post = group['posts'][post_index]
                
                # Размещаем пост frequency раз в неделю
                placed_count = 0
                attempts = 0
                
                while placed_count < frequency and attempts < 50:
                    # Ищем день с минимальной загрузкой
                    day_loads = [(d, get_day_load(week_idx, d))
                                for d in range(7)
                                if len(all_days[week_idx][d]['posts']) < posts_per_day]
                    
                    if not day_loads:
                        break
                    
                    # Сортируем дни по загруженности
                    day_loads.sort(key=lambda x: x[1])
                    
                    # Пробуем разместить пост в один из дней с минимальной загрузкой
                    placed = False
                    for day_idx, _ in day_loads:
                        if can_add_post(all_days, week_idx, day_idx, post['title']):
                            all_days[week_idx][day_idx]['posts'].append({
                                'title': post['title'],
                                'link': post['link']
                            })
                            placed_count += 1
                            post['placed_in_week'][week_idx] += 1
                            placed = True
                            break
                    
                    # Если не удалось разместить пост, увеличиваем счетчик попыток
                    if not placed:
                        attempts += 1
                    
        # Обрабатываем посты без чередования
        for post in weekly_posts:
            if not post['in_alternation_group']:
                # Пропускаем, если интервал 2 недели и это нечетная неделя
                if post.get('interval_weeks', 1) == 2 and week_idx % 2 != 0:
                    continue
                    
                frequency = post['frequency']
                placed_count = 0
                attempts = 0
                
                while placed_count < frequency and attempts < 50:
                    # Ищем день с минимальной загрузкой
                    day_loads = [(d, get_day_load(week_idx, d))
                                for d in range(7)
                                if len(all_days[week_idx][d]['posts']) < posts_per_day]
                    
                    if not day_loads:
                        break
                    
                    # Сортируем дни по загруженности
                    day_loads.sort(key=lambda x: x[1])
                    
                    # Пробуем разместить пост в один из дней с минимальной загрузкой
                    placed = False
                    for day_idx, _ in day_loads:
                        if can_add_post(all_days, week_idx, day_idx, post['title']):
                            all_days[week_idx][day_idx]['posts'].append({
                                'title': post['title'],
                                'link': post['link']
                            })
                            placed_count += 1
                            post['placed_in_week'][week_idx] += 1
                            placed = True
                            break
                    
                    # Если не удалось разместить пост, увеличиваем счетчик попыток
                    if not placed:
                        attempts += 1
    
    # Обрабатываем ежемесячные посты
    for month in range(weeks // 4 + 1):
        month_start_week = month * 4
        month_end_week = min((month + 1) * 4, weeks)
        
        # Обрабатываем группы чередования для ежемесячных постов
        for group_key, group in alternation_groups.items():
            if not group['weekly']:  # Только для ежемесячных групп
                # Пропускаем, если интервал 2 месяца и это нечетный месяц
                if group['interval_months'] == 2 and month % 2 != 0:
                    continue
                
                frequency = group['frequency']
                
                # Определяем, какой пост из группы выходит в этот месяц
                post_index = (month % len(group['posts']))
                post = group['posts'][post_index]
                
                # Размещаем пост frequency раз в месяц
                placed_count = 0
                attempts = 0
                
                while placed_count < frequency and attempts < 50:
                    # Выбираем недели с минимальной загрузкой
                    week_loads = [(w, get_week_load(w)) 
                                for w in range(month_start_week, month_end_week)]
                    
                    if not week_loads:
                        break
                    
                    # Сортируем недели по загруженности
                    week_loads.sort(key=lambda x: x[1])
                    
                    # Пробуем разместить пост в одну из недель с минимальной загрузкой
                    placed = False
                    for week_idx, _ in week_loads:
                        # Ищем день с минимальной загрузкой в этой неделе
                        day_loads = [(d, get_day_load(week_idx, d))
                                    for d in range(7)
                                    if len(all_days[week_idx][d]['posts']) < posts_per_day]
                        
                        if not day_loads:
                            continue
                        
                        # Сортируем дни по загруженности
                        day_loads.sort(key=lambda x: x[1])
                        
                        # Пробуем разместить пост в один из дней с минимальной загрузкой
                        for day_idx, _ in day_loads:
                            if can_add_post(all_days, week_idx, day_idx, post['title'], is_monthly=True):
                                all_days[week_idx][day_idx]['posts'].append({
                                    'title': post['title'],
                                    'link': post['link']
                                })
                                placed_count += 1
                                month_for_post = month // post['interval_months']
                                if month_for_post < len(post['placed_in_month']):
                                    post['placed_in_month'][month_for_post] += 1
                                placed = True
                                break
                        
                        if placed:
                            break
                    
                    # Если не удалось разместить пост, увеличиваем счетчик попыток
                    if not placed:
                        attempts += 1
        
        # Обрабатываем посты без чередования
        for post in monthly_posts:
            if not post['in_alternation_group']:
                # Пропускаем, если интервал 2 месяца и это нечетный месяц
                if post['interval_months'] == 2 and month % 2 != 0:
                    continue
                
                frequency = post['frequency']
                
                # Размещаем пост frequency раз в месяц
                placed_count = 0
                attempts = 0
                
                while placed_count < frequency and attempts < 50:
                    # Выбираем недели с минимальной загрузкой
                    week_loads = [(w, get_week_load(w)) 
                                for w in range(month_start_week, month_end_week)]
                    
                    if not week_loads:
                        break
                    
                    # Сортируем недели по загруженности
                    week_loads.sort(key=lambda x: x[1])
                    
                    # Пробуем разместить пост в одну из недель с минимальной загрузкой
                    placed = False
                    for week_idx, _ in week_loads:
                        # Ищем день с минимальной загрузкой в этой неделе
                        day_loads = [(d, get_day_load(week_idx, d))
                                    for d in range(7)
                                    if len(all_days[week_idx][d]['posts']) < posts_per_day]
                        
                        if not day_loads:
                            continue
                        
                        # Сортируем дни по загруженности
                        day_loads.sort(key=lambda x: x[1])
                        
                        # Пробуем разместить пост в один из дней с минимальной загрузкой
                        for day_idx, _ in day_loads:
                            if can_add_post(all_days, week_idx, day_idx, post['title'], is_monthly=True):
                                all_days[week_idx][day_idx]['posts'].append({
                                    'title': post['title'],
                                    'link': post['link']
                                })
                                placed_count += 1
                                month_for_post = month // post['interval_months']
                                if month_for_post < len(post['placed_in_month']):
                                    post['placed_in_month'][month_for_post] += 1
                                placed = True
                                break
                        
                        if placed:
                            break
                    
                    # Если не удалось разместить пост, увеличиваем счетчик попыток
                    if not placed:
                        attempts += 1
    
    # Проверяем, все ли посты размещены
    for post in weekly_posts:
        if not post['in_alternation_group']:
            for week_idx, placed in enumerate(post['placed_in_week']):
                # Учитываем интервал недель
                if week_idx % post.get('interval_weeks', 1) == 0 and placed < post['frequency']:
                    print(f"Внимание: Пост '{post['title']}' размещен {placed} раз в неделе {week_idx + 1} "
                          f"вместо требуемых {post['frequency']}")
    
    for post in monthly_posts:
        if not post['in_alternation_group']:
            for month_idx, placed in enumerate(post['placed_in_month']):
                if month_idx % post['interval_months'] == 0 and placed < post['frequency']:
                    print(f"Внимание: Пост '{post['title']}' размещен {placed} раз в месяце {month_idx + 1} "
                          f"вместо требуемых {post['frequency']}")
    
    # Балансируем загрузку дней
    for week_idx in range(weeks):
        for day_idx in range(7):
            current_load = get_day_load(week_idx, day_idx)
            if current_load > 2:  # Если в дне больше 2 постов
                adjacent_days = [d for d in range(7) if abs(d - day_idx) == 1]
                for adj_day in adjacent_days:
                    if get_day_load(week_idx, adj_day) == 0:
                        # Пытаемся переместить пост
                        for post in all_days[week_idx][day_idx]['posts']:
                            if post['title'] and can_add_post(all_days, week_idx, adj_day, post['title']):
                                all_days[week_idx][adj_day]['posts'].append(post)
                                all_days[week_idx][day_idx]['posts'].remove(post)
                                break
    
    # Преобразуем в плоский список для вывода
    schedule = []
    for week_days in all_days:
        for day in week_days:
            # Добавляем пустые слоты, если нужно
            while len(day['posts']) < 5:
                day['posts'].append({
                    'title': '',
                    'link': ''
                })
            
            # Добавляем все слоты в расписание
            for i, post in enumerate(day['posts']):
                schedule.append({
                    'day': day['date'],
                    'time': post_slots[i],
                    'title': post['title'],
                    'link': post['link']
                })
    
    return schedule

def create_output_excel(schedule, output_file):
    """Создание выходного Excel файла"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Контент план"
    
    # Заголовки
    headers = ['День недели', 'Дата', 'Время', 'Название поста', 'Ссылка']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Данные
    current_row = 2
    current_day = None
    
    for item in schedule:
        if current_day != item['day']:
            if current_day is not None:
                # Объединяем ячейки для предыдущего дня
                ws.merge_cells(f'A{current_row-5}:A{current_row-1}')
                ws.merge_cells(f'B{current_row-5}:B{current_row-1}')
            
            current_day = item['day']
            # Добавляем пустую строку между днями, если это не первый день
            if current_row > 2:
                current_row += 1
        
        # Записываем данные
        date_cell = ws.cell(row=current_row, column=2)
        date_cell.value = item['day']
        date_cell.number_format = 'dd.mm.yyyy'
        
        weekday_cell = ws.cell(row=current_row, column=1)
        weekday_cell.value = f'=ТЕКСТ(B{current_row};"дддд")'
        
        ws.cell(row=current_row, column=3, value=item['time'])
        ws.cell(row=current_row, column=4, value=item['title'])
        ws.cell(row=current_row, column=5, value=item['link'])
        
        current_row += 1
    
    # Объединяем ячейки для последнего дня
    if current_day is not None:
        ws.merge_cells(f'A{current_row-5}:A{current_row-1}')
        ws.merge_cells(f'B{current_row-5}:B{current_row-1}')
    
    # Форматирование
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Сохранение файла
    wb.save(output_file)

def calculate_recommended_posts_per_day(df, weeks):
    """Рассчитывает рекомендуемое количество постов в день"""
    
    # Создание группы чередования для учета в расчетах
    alternation_groups = {}
    
    # Подсчитываем общее количество постов в неделю
    weekly_posts_count = 0
    
    # Подсчитываем общее количество постов в месяц
    monthly_posts_count = 0
    
    # Первый проход: определяем группы чередования
    for idx, row in df.iterrows():
        row_number = idx + 2
        
        # Если у поста есть указание на чередование
        if pd.notna(row['E']):
            alternate_rows = [int(x.strip()) for x in str(row['E']).split(',')]
            group = [row_number] + alternate_rows
            group_key = tuple(sorted(group))
            
            if group_key not in alternation_groups:
                alternation_groups[group_key] = {
                    'weekly': pd.notna(row['B']),
                    'frequency': float(row['B']) if pd.notna(row['B']) else float(row['C']),
                    'count': len(group)
                }
    
    # Второй проход: подсчитываем посты
    for idx, row in df.iterrows():
        row_number = idx + 2
        
        # Проверяем, входит ли пост в группу чередования
        in_alternation_group = False
        for group_key in alternation_groups:
            if row_number in group_key:
                in_alternation_group = True
                break
        
        # Если пост входит в группу чередования, мы уже его учли
        if in_alternation_group:
            continue
        
        # Учитываем обычные посты
        if pd.notna(row['B']):  # Посты в неделю
            frequency = float(row['B'])
            weekly_posts_count += frequency
        elif pd.notna(row['C']):  # Посты в месяц
            frequency = float(row['C'])
            monthly_posts_count += frequency
    
    # Добавляем посты из групп чередования
    for group in alternation_groups.values():
        if group['weekly']:
            weekly_posts_count += group['frequency']
        else:
            monthly_posts_count += group['frequency']
    
    # Рассчитываем общее количество постов за период
    total_posts = weekly_posts_count * weeks + monthly_posts_count * (weeks // 4)
    
    # Рассчитываем общее количество дней в периоде
    total_days = weeks * 7
    
    # Рассчитываем среднее количество постов в день и округляем вверх
    recommended_posts_per_day = int(total_posts / total_days) + 1
    
    # Проверяем, нужно ли увеличить для обеспечения всех условий
    # (например, если много постов должны выходить в одни и те же дни)
    max_posts_per_week = max(weekly_posts_count, monthly_posts_count / 4)
    max_posts_per_day_weekly = int(max_posts_per_week / 3) + 1  # Предполагаем распределение на ~3 дня в неделю
    
    # Берем максимальное из двух значений для гарантии
    recommended_posts_per_day = max(recommended_posts_per_day, max_posts_per_day_weekly)
    
    return recommended_posts_per_day

def main():
    print("Добро пожаловать в планировщик контента!")
    print("\nВведите путь к Excel файлу с данными.")
    print("Примеры путей:")
    print("- Для Windows: C:\\Users\\Username\\Desktop\\posts.xlsx")
    print("- Для Mac/Linux: /Users/Username/Desktop/posts.xlsx")
    print("- Если файл в той же папке, просто введите имя файла: posts.xlsx")
    print("\nФормат Excel файла:")
    print("1-й столбец: Названия постов")
    print("2-й столбец: Количество постов в неделю")
    print("3-й столбец: Количество постов в месяц")
    print("4-й столбец: Ссылка на пост")
    print("5-й столбец: Чередование (номера строк через запятую)")
    
    # Ввод пути к файлу
    while True:
        file_path = input("\nВведите путь к файлу: ").strip()
        if os.path.exists(file_path):
            if file_path.lower().endswith(('.xlsx', '.xls')):
                break
            print("Файл должен быть в формате Excel (.xlsx или .xls)")
        else:
            print("Файл не найден. Проверьте путь и попробуйте еще раз.")
    
    # Чтение файла
    df = read_input_file(file_path)
    if df is None:
        return
    
    # Проверка данных
    is_valid, error_message = validate_input(df)
    if not is_valid:
        print(f"Ошибка в данных: {error_message}")
        return
    
    # Ввод количества недель для планирования
    while True:
        try:
            weeks = int(input("Введите количество недель для планирования: "))
            if weeks > 0:
                break
            print("Количество недель должно быть положительным числом.")
        except ValueError:
            print("Пожалуйста, введите число.")
    
    # Расчет рекомендуемого количества постов в день
    recommended_posts = calculate_recommended_posts_per_day(df, weeks)
    
    # Ввод параметров
    while True:
        try:
            print(f"\nРекомендуемое количество постов в день: {recommended_posts}")
            posts_per_day = int(input("Введите максимальное количество постов в день: "))
            if posts_per_day > 0:
                break
            print("Количество постов должно быть положительным числом.")
        except ValueError:
            print("Пожалуйста, введите число.")
    
    # Ввод начальной даты
    while True:
        try:
            date_str = input("Введите начальную дату в формате ДД.ММ.ГГГГ (или нажмите Enter для текущей даты): ").strip()
            if not date_str:
                start_date = None
                break
            start_date = datetime.strptime(date_str, '%d.%m.%Y')
            if start_date.date() < datetime.now().date():
                print("Начальная дата не может быть в прошлом.")
                continue
            break
        except ValueError:
            print("Неверный формат даты. Используйте формат ДД.ММ.ГГГГ")
    
    # Создание расписания
    schedule = get_post_schedule(df, weeks, posts_per_day, start_date)
    
    # Создание выходного файла
    output_file = "content_plan.xlsx"
    create_output_excel(schedule, output_file)
    print(f"Контент план успешно создан и сохранен в файл {output_file}")

if __name__ == "__main__":
    main() 