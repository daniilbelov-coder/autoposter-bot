"""
Модуль для генерации еженедельного календаря коммуникаций в Excel.
"""

import os
from datetime import datetime, timedelta
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import POSTING_TIMES, TIMEZONE, logger
from message_selector import selector
from database import db


def get_week_dates(start_date: datetime = None) -> List[datetime]:
    """
    Получить даты недели (понедельник-воскресенье).
    
    Args:
        start_date: Начальная дата (если None - текущий понедельник)
        
    Returns:
        Список дат недели
    """
    if start_date is None:
        today = datetime.now(TIMEZONE)
        # Получить понедельник текущей недели
        start_date = today - timedelta(days=today.weekday())
    
    week_dates = []
    for i in range(7):
        week_dates.append(start_date + timedelta(days=i))
    
    return week_dates


def get_planned_posts_for_week() -> List[Dict]:
    """
    Получить список запланированных постов на неделю.
    
    Returns:
        Список словарей с информацией о постах
    """
    week_dates = get_week_dates()
    planned_posts = []
    
    # Для каждого дня недели
    for date in week_dates:
        # Для каждого времени отправки
        for time_config in POSTING_TIMES:
            hour = time_config['hour']
            minute = time_config['minute']
            
            post_datetime = date.replace(hour=hour, minute=minute, second=0, microsecond=0)
            
            # Получить доступные сообщения (упрощенная версия - берем все)
            available_messages = selector.messages
            
            if available_messages:
                # Случайное сообщение для примера
                import random
                message = random.choice(available_messages)
                
                planned_posts.append({
                    'date': post_datetime.date(),
                    'time': post_datetime.time(),
                    'datetime': post_datetime,
                    'day_name': post_datetime.strftime('%A'),
                    'message_id': message['id'],
                    'message_title': message['title'],
                    'frequency': message['frequency'],
                    'has_photos': len(message.get('photos', [])) > 0,
                    'photos_count': len(message.get('photos', []))
                })
    
    # Сортировать по дате и времени
    planned_posts.sort(key=lambda x: x['datetime'])
    
    return planned_posts


def create_schedule_excel(output_path: str = "schedule.xlsx") -> str:
    """
    Создать Excel файл с календарем коммуникаций.
    
    Args:
        output_path: Путь для сохранения файла
        
    Returns:
        Путь к созданному файлу
    """
    logger.info("Начало генерации календаря коммуникаций")
    
    # Создать книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Календарь коммуникаций"
    
    # Получить даты недели
    week_dates = get_week_dates()
    start_date = week_dates[0]
    end_date = week_dates[-1]
    
    # Получить запланированные посты
    planned_posts = get_planned_posts_for_week()
    
    # Стили
    title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='Arial', size=10)
    
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    alternate_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Заголовок документа
    ws.merge_cells('A1:G1')
    ws['A1'] = f'КАЛЕНДАРЬ КОММУНИКАЦИЙ'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 25
    
    # Подзаголовок с датами
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Период: {start_date.strftime("%d.%m.%Y")} - {end_date.strftime("%d.%m.%Y")}'
    ws['A2'].font = Font(name='Arial', size=11, italic=True)
    ws['A2'].alignment = center_alignment
    ws.row_dimensions[2].height = 20
    
    # Пустая строка
    ws.row_dimensions[3].height = 5
    
    # Заголовки столбцов
    headers = ['№', 'Дата', 'День недели', 'Время', 'Название сообщения', 'Частота', 'Фото']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    ws.row_dimensions[4].height = 20
    
    # Данные
    day_names_ru = {
        'Monday': 'Понедельник',
        'Tuesday': 'Вторник',
        'Wednesday': 'Среда',
        'Thursday': 'Четверг',
        'Friday': 'Пятница',
        'Saturday': 'Суббота',
        'Sunday': 'Воскресенье'
    }
    
    for idx, post in enumerate(planned_posts, 1):
        row = idx + 4
        
        # Чередующийся цвет строк
        fill = alternate_fill if idx % 2 == 0 else PatternFill()
        
        # Номер
        cell = ws.cell(row=row, column=1, value=idx)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Дата
        cell = ws.cell(row=row, column=2, value=post['date'].strftime('%d.%m.%Y'))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # День недели
        day_name = day_names_ru.get(post['day_name'], post['day_name'])
        cell = ws.cell(row=row, column=3, value=day_name)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Время
        cell = ws.cell(row=row, column=4, value=post['time'].strftime('%H:%M'))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Название сообщения
        cell = ws.cell(row=row, column=5, value=post['message_title'])
        cell.font = cell_font
        cell.alignment = left_alignment
        cell.border = border
        cell.fill = fill
        
        # Частота
        frequency_ru = {
            'daily': 'Ежедневно',
            'weekly': 'Еженедельно',
            'biweekly': 'Раз в 2 недели',
            'monthly': 'Ежемесячно'
        }
        freq_text = frequency_ru.get(post['frequency'], post['frequency'])
        cell = ws.cell(row=row, column=6, value=freq_text)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        # Фото
        photos_text = f"✓ ({post['photos_count']})" if post['has_photos'] else "—"
        cell = ws.cell(row=row, column=7, value=photos_text)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = fill
        
        ws.row_dimensions[row].height = 18
    
    # Футер
    footer_row = len(planned_posts) + 6
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    ws[f'A{footer_row}'] = f'Всего запланировано публикаций: {len(planned_posts)}'
    ws[f'A{footer_row}'].font = Font(name='Arial', size=10, italic=True)
    ws[f'A{footer_row}'].alignment = center_alignment
    
    # Подпись
    footer_row += 1
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    ws[f'A{footer_row}'] = f'Сгенерировано: {datetime.now(TIMEZONE).strftime("%d.%m.%Y %H:%M")}'
    ws[f'A{footer_row}'].font = Font(name='Arial', size=9, italic=True, color='808080')
    ws[f'A{footer_row}'].alignment = center_alignment
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10
    
    # Сохранить файл
    wb.save(output_path)
    logger.info(f"Календарь сохранён: {output_path}")
    
    return output_path


def generate_schedule_for_subscribers() -> str:
    """
    Сгенерировать календарь для рассылки подписчикам.
    
    Returns:
        Путь к файлу
    """
    # Создать папку для расписаний, если нет
    schedules_dir = "schedules"
    if not os.path.exists(schedules_dir):
        os.makedirs(schedules_dir)
    
    # Имя файла с датой
    filename = f"calendar_{datetime.now(TIMEZONE).strftime('%Y_%m_%d')}.xlsx"
    filepath = os.path.join(schedules_dir, filename)
    
    return create_schedule_excel(filepath)


if __name__ == '__main__':
    # Тестовая генерация
    print("Генерация тестового календаря...")
    filepath = generate_schedule_for_subscribers()
    print(f"✅ Календарь создан: {filepath}")
